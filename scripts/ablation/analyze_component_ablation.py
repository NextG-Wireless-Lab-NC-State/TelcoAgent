"""Component-ablation analysis — paired statistics + summary tables.

Consumes the two upstream artefacts produced by the component-ablation
pipeline and emits the per-axis statistics that the REPORT.md cites:

Inputs
------
* ``judge/judge_per_station.csv`` — RAGAS Faithfulness + Answer
  Relevancy per (station, arm), arms = kg_on / kg_off / minus_predictor
  / minus_explainer (from ``run_explainer_ablation_judge.py``).
* ``prediction_per_station.csv`` — Chronos-2 vs seasonal-naive nRMSE /
  MASE per station (from ``compute_prediction_metrics.py``).

Comparisons (all Full-referenced, station-paired)
-------------------------------------------------
Explanation — Full(kg_on) vs each removal arm, on Faithfulness and AR:
* −KG          : kg_off vs kg_on
* −Predictor   : minus_predictor vs kg_on
* −Explainer   : minus_explainer vs kg_on
Prediction — Chronos-2 vs seasonal-naive on mean_nRMSE and mean_MASE.

Each comparison reports mean_diff, 95 % bootstrap CI, Wilcoxon p-value,
paired Cohen's d, and n — straight from ``stats.paired_comparison``.

Faithfulness is scored on all 30 stations. Answer Relevancy is only
meaningful on event-bearing stations (AR = 0 when no anomaly events are
detected); the AR comparison is therefore restricted to stations where
**both** paired arms have at least one scored event.

Outputs (under ``--out-dir``)
-----------------------------
* ``explanation_per_station.csv`` — wide table, one row per station.
* ``component_stats.csv`` / ``.json`` — one row per comparison.
* ``ablation_summary.xlsx`` — workbook: prediction / explanation /
  stats / per-station sheets.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path
from typing import Dict, List, Optional

import numpy as np

_REPO_ROOT = Path(__file__).resolve().parents[2]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from scripts.ablation.stats import paired_comparison  # noqa: E402

_ARMS = ("kg_on", "kg_off", "minus_predictor", "minus_explainer")
_KPI_NAMES = (
    "RRC_Conn",
    "DL_CQI",
    "DL_iBler",
    "DL_rBler",
    "MAC_DL_Eff",
    "PRB_Util",
    "Throughput",
)


def _norm_station(name: str) -> str:
    return name[len("station_") :] if name.startswith("station_") else name


def _load_judge(judge_csv: Path) -> Dict[str, Dict[str, Dict[str, Optional[float]]]]:
    """Return {station -> {arm -> {faithfulness, answer_relevancy, n_events}}}."""
    out: Dict[str, Dict[str, Dict[str, Optional[float]]]] = {}
    with judge_csv.open() as fh:
        for r in csv.DictReader(fh):
            sid = _norm_station(r["station_id"])
            arm = r["arm"]

            def _f(key: str) -> Optional[float]:
                raw = r.get(key, "")
                if raw is None or raw == "":
                    return None
                try:
                    return float(raw)
                except ValueError:
                    return None

            out.setdefault(sid, {})[arm] = {
                "faithfulness": _f("faithfulness"),
                "answer_relevancy": _f("answer_relevancy"),
                "n_events": _f("n_events"),
                "n_scored": _f("n_scored"),
            }
    return out


def _load_prediction(pred_csv: Path) -> Dict[str, Dict[str, Dict[str, float]]]:
    """Return {station -> {forecaster -> {metric -> value}}}."""
    out: Dict[str, Dict[str, Dict[str, float]]] = {}
    with pred_csv.open() as fh:
        for r in csv.DictReader(fh):
            sid = _norm_station(r["station_id"])
            out.setdefault(sid, {})[r["forecaster"]] = {
                k: float(v) for k, v in r.items() if k not in ("station_id", "forecaster")
            }
    return out


def _paired_arrays(
    judge: Dict, stations: List[str], arm_a: str, arm_b: str, metric: str, drop_zero_ar: bool
):
    """Return (arr_a, arr_b, used_stations) for a paired arm comparison."""
    arr_a: List[float] = []
    arr_b: List[float] = []
    used: List[str] = []
    for sid in stations:
        rec = judge.get(sid, {})
        va = rec.get(arm_a, {}).get(metric)
        vb = rec.get(arm_b, {}).get(metric)
        if va is None or vb is None:
            continue
        if drop_zero_ar:
            # AR is only meaningful where both arms scored >=1 event.
            ea = rec.get(arm_a, {}).get("n_events") or 0.0
            eb = rec.get(arm_b, {}).get("n_events") or 0.0
            if ea < 1 or eb < 1:
                continue
        arr_a.append(va)
        arr_b.append(vb)
        used.append(sid)
    return np.array(arr_a), np.array(arr_b), used


def _comparison_row(label: str, axis: str, metric: str, arr_a, arr_b) -> Dict:
    """Build a stats row; arr_a = removal arm, arr_b = Full (reference)."""
    if len(arr_a) < 2:
        return {
            "comparison": label,
            "axis": axis,
            "metric": metric,
            "n": len(arr_a),
            "mean_removal": round(float(np.mean(arr_a)), 6) if len(arr_a) else float("nan"),
            "mean_full": round(float(np.mean(arr_b)), 6) if len(arr_b) else float("nan"),
            "mean_diff": float("nan"),
            "ci_low": float("nan"),
            "ci_high": float("nan"),
            "p_value": float("nan"),
            "cohens_d": float("nan"),
        }
    res = paired_comparison(arr_a, arr_b)
    return {
        "comparison": label,
        "axis": axis,
        "metric": metric,
        "n": res["n"],
        "mean_removal": round(float(np.mean(arr_a)), 6),
        "mean_full": round(float(np.mean(arr_b)), 6),
        "mean_diff": round(res["mean_diff"], 6),
        "ci_low": round(res["ci_low"], 6),
        "ci_high": round(res["ci_high"], 6),
        "p_value": round(res["p_value"], 6),
        "cohens_d": round(res["cohens_d"], 6),
    }


def main(argv: List[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Component-ablation paired statistics + summary tables."
    )
    parser.add_argument("--judge-csv", required=True)
    parser.add_argument("--prediction-csv", required=True)
    parser.add_argument("--stations-file", required=True)
    parser.add_argument("--out-dir", required=True)
    args = parser.parse_args(argv)

    stations = [
        _norm_station(line.strip())
        for line in Path(args.stations_file).read_text().splitlines()
        if line.strip()
    ]
    judge = _load_judge(Path(args.judge_csv))
    pred = _load_prediction(Path(args.prediction_csv))
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # ── per-station explanation table ────────────────────────────────────
    expl_rows: List[Dict] = []
    for sid in stations:
        rec = judge.get(sid, {})
        row: Dict[str, object] = {"station_id": sid}
        for arm in _ARMS:
            a = rec.get(arm, {})
            row[f"{arm}_faithfulness"] = a.get("faithfulness")
            row[f"{arm}_answer_relevancy"] = a.get("answer_relevancy")
            row[f"{arm}_n_events"] = a.get("n_events")
        expl_rows.append(row)
    expl_path = out_dir / "explanation_per_station.csv"
    with expl_path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(expl_rows[0].keys()))
        writer.writeheader()
        writer.writerows(expl_rows)

    # ── stats: explanation comparisons (Full = kg_on reference) ──────────
    stat_rows: List[Dict] = []
    removal_arms = [
        ("-KG", "kg_off"),
        ("-Predictor", "minus_predictor"),
        ("-Explainer", "minus_explainer"),
    ]
    for label, arm in removal_arms:
        for metric in ("faithfulness", "answer_relevancy"):
            drop_zero = metric == "answer_relevancy"
            arr_a, arr_b, used = _paired_arrays(judge, stations, arm, "kg_on", metric, drop_zero)
            stat_rows.append(
                _comparison_row(f"{label} vs Full", "explanation", metric, arr_a, arr_b)
            )

    # ── stats: prediction comparison (Chronos-2 vs seasonal-naive) ───────
    for metric in ("mean_nRMSE", "mean_MASE"):
        chronos = np.array([pred[s]["chronos2"][metric] for s in stations if s in pred])
        naive = np.array([pred[s]["seasonal_naive"][metric] for s in stations if s in pred])
        # removal arm = naive (-Predictor), reference = chronos2 (Full)
        stat_rows.append(
            _comparison_row("-Predictor vs Full", "prediction", metric, naive, chronos)
        )

    stats_path = out_dir / "component_stats.csv"
    fieldnames = [
        "comparison",
        "axis",
        "metric",
        "n",
        "mean_full",
        "mean_removal",
        "mean_diff",
        "ci_low",
        "ci_high",
        "p_value",
        "cohens_d",
    ]
    with stats_path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(stat_rows)
    (out_dir / "component_stats.json").write_text(json.dumps(stat_rows, indent=2), encoding="utf-8")

    # ── integrated workbook ──────────────────────────────────────────────
    _write_xlsx(out_dir, stat_rows, expl_rows, pred, stations)

    print(f"Wrote {expl_path}")
    print(f"Wrote {stats_path}")
    print(f"Wrote {out_dir / 'ablation_summary.xlsx'}")
    print("\n=== component statistics ===")
    for r in stat_rows:
        print(
            f"  {r['comparison']:22s} {r['axis']:11s} {r['metric']:18s} "
            f"diff={r['mean_diff']!s:>10}  "
            f"CI=[{r['ci_low']}, {r['ci_high']}]  "
            f"p={r['p_value']}  d={r['cohens_d']}  n={r['n']}"
        )
    return 0


def _write_xlsx(out_dir: Path, stat_rows, expl_rows, pred, stations) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    # prediction sheet
    ws = wb.create_sheet("prediction")
    ws.append(["station_id", "forecaster", "mean_nRMSE", "mean_MASE", "overall_nRMSE"])
    for sid in stations:
        if sid not in pred:
            continue
        for fc in ("chronos2", "seasonal_naive"):
            m = pred[sid][fc]
            ws.append([sid, fc, m["mean_nRMSE"], m["mean_MASE"], m["overall_nRMSE"]])

    # explanation sheet
    ws = wb.create_sheet("explanation")
    headers = list(expl_rows[0].keys())
    ws.append(headers)
    for row in expl_rows:
        ws.append([row.get(h) for h in headers])

    # stats sheet
    ws = wb.create_sheet("stats")
    headers = list(stat_rows[0].keys())
    ws.append(headers)
    for row in stat_rows:
        ws.append([row.get(h) for h in headers])

    wb.save(out_dir / "ablation_summary.xlsx")


if __name__ == "__main__":
    raise SystemExit(main())
