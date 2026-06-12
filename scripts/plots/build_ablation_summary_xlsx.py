#!/usr/bin/env python3
"""Build integrated workbook output/ablation/ablation_summary.xlsx.

Consumes:
  - output/ablation/refinement_onoff/results.csv
  - output/ablation/kg_onoff/results.csv
  - output/ablation/context_sweep/results.csv
  - output/ablation/qth_sweep/results.csv (may be empty if AX-4 skipped)

Writes one sheet per axis + a "summary" sheet with cost.json contents.

Idempotent.
"""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

ROOT = Path("/home/gkim26/Desktop/workplace/telcoagent/output/ablation")


def _read_csv(p: Path):
    if not p.exists():
        return []
    with open(p) as f:
        return list(csv.DictReader(f))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", type=Path, default=ROOT / "ablation_summary.xlsx")
    args = ap.parse_args()

    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    axes = [
        ("AX-1_refinement_onoff", ROOT / "refinement_onoff" / "results.csv"),
        ("AX-2_kg_onoff", ROOT / "kg_onoff" / "results.csv"),
        ("AX-3_context_sweep", ROOT / "context_sweep" / "results.csv"),
        ("AX-4_qth_sweep", ROOT / "qth_sweep" / "results.csv"),
    ]
    for sheet_name, csv_path in axes:
        ws = wb.create_sheet(sheet_name[:31])
        rows = _read_csv(csv_path)
        if rows:
            ws.append(list(rows[0].keys()))
            for r in rows:
                ws.append(list(r.values()))
        else:
            ws.append(["status"])
            ws.append(["EMPTY / SKIPPED — see REPORT.md"])

    # Summary sheet: cost rollup
    ws = wb.create_sheet("summary")
    ws.append(["axis", "tokens_in", "tokens_out", "usd_actual", "wallclock_sec"])
    for axis_dir in ("refinement_onoff", "kg_onoff", "context_sweep", "qth_sweep"):
        cj = ROOT / axis_dir / "cost.json"
        if cj.exists():
            d = json.load(open(cj))
            ws.append(
                [
                    axis_dir,
                    d.get("tokens_in"),
                    d.get("tokens_out"),
                    d.get("usd_actual"),
                    d.get("wallclock_sec"),
                ]
            )
        else:
            ws.append([axis_dir, None, None, None, None])

    wb.save(args.out)
    print(f"wrote {args.out}")


if __name__ == "__main__":
    main()
